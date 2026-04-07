import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import time
from urllib.parse import urlparse
import threading
import os
from openpyxl.styles import numbers
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import webbrowser
import sys

def caminho_recurso(rel_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, rel_path)

# --- Funções existentes (simplificadas) ---
def extrair_fornecedor(url):
    dominio = urlparse(url).netloc.replace("www.", "")
    partes = dominio.split(".")
    return partes[-3] if len(partes) >= 3 else partes[0]

def limpar_preco(texto):
    return texto.replace("\xa0", " ").strip()

def extrair_preco(driver, fornecedor):
    try:
        # --- KaBuM ---
        if "kabum" in fornecedor:
            el = driver.find_element(By.XPATH, "//h4[contains(@class,'text-4xl')]")
            return limpar_preco(el.text)
        # --- Mercado Livre ---
        elif "mercadolivre" in fornecedor:
            el = driver.find_element(By.XPATH, "//meta[@itemprop='price']")
            preco = el.get_attribute("content")
            return f"R$ {float(preco):.2f}".replace(".", ",")
        # --- DetonaShop ---
        elif "detonashop" in fornecedor:
            el = driver.find_element(By.XPATH, "//span[contains(@id, 'product-price')]")
            preco = el.get_attribute("data-price-amount")
            return f"R$ {float(preco):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        # --- fallback ---
        else:
            return "Não configurado"
    except Exception as e:
        print(f"Erro ao extrair preço ({fornecedor}): {e}")
        return "Não encontrado"

def processar_entrada(linhas):
    dados = []
    for linha in linhas:
        partes = linha.strip().split()
        if not partes:
            continue
        url = partes[0]
        # Quantidade padrão = 1
        quantidade = 1
        if len(partes) > 1:
            try:
                quantidade = max(1, int(partes[1]))
            except:
                pass
        dados.append((url, quantidade))
    return dados

# --- Função principal (em thread) ---
def gerar_planilha_thread():
    texto = caixa_links.get("1.0", tk.END)
    linhas = texto.split("\n")
    entrada = processar_entrada(linhas)
    if not entrada:
        messagebox.showwarning("Aviso", "Nenhum link válido informado.")
        return
    # Navegador em 2º plano
    options = webdriver.ChromeOptions()
    options.add_argument("--start-minimized")
    driver = webdriver.Chrome(options=options)
    driver.minimize_window()
    dados = []
    total_links = len(entrada)
    for i, (url, quantidade) in enumerate(entrada):
        driver.get(url)
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "h1")))
            try:
                nome = driver.find_element(By.TAG_NAME, "h1").text
                fornecedor = extrair_fornecedor(url)
                preco = extrair_preco(driver, fornecedor)
                dados.append({
                    "Item": nome,
                    "Valor unitário": preco,
                    "Qtde": quantidade,
                    "Valor total": "",
                    "Fornecedor": fornecedor,
                    "Links": url
                })
            except Exception as e:
                print("Erro:", e)
        except TimeoutException:
            print(f"[WARN] Timeout carregando página: {url}")
        except Exception as e:
            print("Erro:", e)
        progresso["value"] = ((i + 1) / total_links) * 100
        janela.update_idletasks()
    driver.quit()
    df = pd.DataFrame(dados)
    # Fórmula linha a linha
    for i in range(len(df)):
        df.loc[i, "Valor total"] = f"=B{i+2}*C{i+2}"
    caminho = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Salvar planilha como"
    )
    if not caminho:
        return
    # Salvar Excel
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Planilha1")
        ws = writer.sheets["Planilha1"]
        ultima_linha = len(df) + 2
        ws[f"D{ultima_linha}"] = f"=SUM(D2:D{ultima_linha-1})"
        for i in range(2, ultima_linha):
            ws[f"D{i}"].number_format = 'R$ #,##0.00'
        ws[f"D{ultima_linha}"].number_format = 'R$ #,##0.00'
    messagebox.showinfo("Sucesso", "Planilha criada!")
    janela.after(0, lambda: botao_gerar.config(state="normal"))
    janela.after(0, lambda: caixa_links.delete("1.0", tk.END))
    janela.after(0, lambda: progresso.config(value=0))
    # Abrir automaticamente
    if abrir_var.get():
        os.startfile(caminho)

# Wrapper p/ thread
def gerar_planilha():
    botao_gerar.config(state="disabled")
    threading.Thread(target=gerar_planilha_thread).start()

def mostrar_sobre():
    janela_sobre = tk.Toplevel()
    janela_sobre.title("Sobre o ESGen")
    janela_sobre.geometry("350x200")

    tk.Label(janela_sobre, text="ESGen - Excel Sheet Generator", font=("Arial", 12, "bold")).pack(pady=5)
    tk.Label(janela_sobre, text="Versão: 1.0.0 (beta)").pack()
    tk.Label(janela_sobre, text="Autor: Timóteo A. B. da Silva").pack(pady=5)

    link_github = tk.Label(janela_sobre, text="Repositório GitHub", fg="blue", cursor="hand2")
    link_github.pack()
    link_github.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/imbaTIMvel/excel_sheet_generator/tree/main"))

    link_manual = tk.Label(janela_sobre, text="Manual do usuário", fg="blue", cursor="hand2")
    link_manual.pack(pady=5)
    link_manual.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/imbaTIMvel/excel_sheet_generator/tree/main/docs/esgen-user_manual-v1.0.0.pdf"))

    tk.Button(janela_sobre, text="Fechar", command=janela_sobre.destroy).pack(pady=10)

# --- UI ---
janela = tk.Tk()
janela.title("ESGen")
janela.iconbitmap(caminho_recurso("assets/icon.ico"))

menu_bar = tk.Menu(janela)
janela.config(menu=menu_bar)
menu_ajuda = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Ajuda", menu=menu_ajuda)
menu_ajuda.add_command(label="Sobre", command=mostrar_sobre)

tk.Label(janela, text="Cole os links:").pack()

caixa_links = tk.Text(janela, height=15, width=60)
caixa_links.pack()

abrir_var = tk.BooleanVar()
tk.Checkbutton(janela, text="Abrir planilha quando estiver pronta", variable=abrir_var).pack()

progresso = ttk.Progressbar(janela, orient="horizontal", length=400, mode="determinate")
progresso.pack(pady=10)

botao_gerar = tk.Button(janela, text="Gerar Planilha", command=gerar_planilha)
botao_gerar.pack(pady=10)

janela.mainloop()
